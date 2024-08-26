class Json():

    dictionary = None

    def __init__(self, dictionary:dict):
        self.dictionary = dictionary
        for atribute in dictionary:
            setattr(self, atribute, dictionary[atribute])
    
    def getNewDictionary(self, parameters:list=[]):
        newDict = {}
        if parameters != []:
        
            for parameter in parameters:
                param = getattr(self, parameter)
                newDict[parameter] = param
        else:
            newDict = self.buildClass()
                
        return newDict

    def buildClass(self):
        return self.__dict__
    
    def add(self, name:str, value:any):
        setattr(self, name, value)


import openpyxl
from openpyxl.utils import get_column_letter
# Crear un nuevo libro de Excel

class ExcelGenerator:
    def __init__(self, registers:list):
        self.registers = registers

    def generate(self):
            libro = openpyxl.Workbook()

            # Obtener la hoja activa (por defecto, será la primera hoja)
            sheet = libro.active
            headers = [
                "Archivo",
                "Usuario",
                "Fecha y Hora",
                "Acción"
            ]
            
            # Escribir los encabezados en la primera fila
            sheet.append(headers)
            sheet.title = "Historial"

            # Llenar el archivo con los datos de las polizas
            for register in self.registers:
                data = Json(register)
                added = [
                    data.name.split(" - ")[0],
                    data.name.split("- Usuario: ")[-1],
                    data.date,
                    data.action
                ]
                sheet.append(added)
                
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        # Obtener el largo del contenido de la celda
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Ajustar el ancho de la columna en base al contenido más largo
                adjusted_width = max_length + 2  # Añadir un pequeño margen
                sheet.column_dimensions[column_letter].width = adjusted_width
                
            route = 'files/history.xlsx'

            # Guardar los datos en el archivo de Excel en la ruta seleccionada
            if route != "":
                libro.save(route)
                return route